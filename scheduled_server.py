
#!/usr/bin/env python3
# -*- coding: utf-8 -*-

# --- CLEAN IMPORTS AT TOP ---

import sys
import os
import subprocess
import json
from datetime import datetime
from flask import Flask, jsonify, send_from_directory, make_response, Response
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger

app = Flask(__name__)
scheduler = BackgroundScheduler()

scheduler = BackgroundScheduler()





# Venv path'i ekle
venv_site_packages = os.path.join(os.path.dirname(__file__), '.venv', 'Lib', 'site-packages')
if os.path.exists(venv_site_packages):
    sys.path.insert(0, venv_site_packages)

def excel_to_json():
    """Convert database.xlsx to database.json using openpyxl"""
    try:
        from openpyxl import load_workbook
        
        xlsx_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'database.xlsx')
        json_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'database.json')
        
        if not os.path.exists(xlsx_file):
            return False
        
        # Read Excel with openpyxl
        wb = load_workbook(xlsx_file)
        ws = wb['Sonuçlar']
        
        # Get headers from first row
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)
        
        # Get data rows
        records = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            record = {}
            for col_idx, value in enumerate(row):
                if col_idx < len(headers):
                    record[headers[col_idx]] = value
            records.append(record)
        
        # Write to JSON (compact format) - UTF-8 BOM-less
        with open(json_file, 'w', encoding='utf-8-sig') as f:
            json.dump(records, f, ensure_ascii=False, indent=None)
        
        return True
    except Exception as e:
        print(f"[ERROR] excel_to_json: {e}")
        return False

app = Flask(__name__, static_folder=os.path.dirname(os.path.abspath(__file__)), static_url_path='')
scheduler = BackgroundScheduler()

# Cache kontrol - dosyalar cache'lenmesin
@app.after_request
def set_cache_control(response):
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response

# Scheduled job function
def run_daily_update():
    """Her gün 23:59'da hosgoru_takvim_bot.py'ı çalıştır"""
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"\n{'='*60}")
    print(f"[{timestamp}] OTOMATİK GÜNCELLEME BAŞLANDI")
    print(f"{'='*60}")
    
    try:
        # Bot'ı çalıştır
        result = subprocess.run(
            ['python', 'hosgoru_takvim_bot.py', '--lang', 'tr', '--retries', '5'],
            cwd=os.path.dirname(os.path.abspath(__file__)),
            capture_output=True,
            text=True,
            timeout=3600  # Max 1 saat
        )
        
        if result.returncode == 0:
            print(f"[{timestamp}] ✓ GÜNCELLEME BAŞARILI")
            # Excel'i JSON'a çevir
            if excel_to_json():
                print(f"[{timestamp}] ✓ database.json güncelendi")
            else:
                print(f"[{timestamp}] ✗ database.json güncellenemedi")
        else:
            print(f"[{timestamp}] ✗ GÜNCELLEME BAŞARISIZ")
            print(f"Hata: {result.stderr}")
            
    except subprocess.TimeoutExpired:
        print(f"[{timestamp}] ✗ ZAMAN AŞIMI (1 saat)")
    except Exception as e:
        print(f"[{timestamp}] ✗ HATA: {str(e)}")

@app.route('/')
def home():
    return send_from_directory(os.path.dirname(os.path.abspath(__file__)), 'index.html')

@app.route('/status')
def status():
    """Scheduler durumunu göster"""
    jobs = scheduler.get_jobs()
    return jsonify({
        "scheduler_active": scheduler.running,
        "scheduled_jobs": [
            {
                "name": job.name,
                "trigger": str(job.trigger),
                "next_run": str(job.next_run_time)
            }
            for job in jobs
        ]
    })


# --- FIXED ENCODING & ADDED /get_database_temp ---
def _send_json_file(filename):
    try:
        json_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), filename)
        print(f"[DEBUG] Reading file: {json_file}")
        if not os.path.exists(json_file):
            print(f"[DEBUG] File not found: {json_file}")
            return jsonify({"error": f"{filename} not found"}), 404
        with open(json_file, 'r', encoding='utf-8-sig') as f:
            file_content = f.read()
            print(f"[DEBUG] File content (first 500 chars): {file_content[:500]}")
            f.seek(0)
            data = json.load(f)
        return jsonify(data)
    except Exception as e:
        import traceback
        print(f"❌ ERROR in /{filename}: {str(e)}")
        print(traceback.format_exc())
        return jsonify({"error": str(e)}), 500

@app.route('/get_database')
def get_database():
    """database.json dosyasını JSON olarak döndür"""
    print(f"[/get_database] Sending database.json...")
    return _send_json_file('database.json')

@app.route('/get_database_temp')
def get_database_temp():
    """database_temp.json dosyasını JSON olarak döndür"""
    print(f"[/get_database_temp] Sending database_temp.json...")
    return _send_json_file('database_temp.json')

@app.route('/update-now', methods=['POST'])
def update_now():
    """Acil güncelleme tetikle"""
    print("[MANUEL] Acil güncelleme tetikleniyor...")
    run_daily_update()
    return jsonify({"status": "güncelleme başlatıldı"})

@app.route('/<filename>')
def serve_static(filename):
    return send_from_directory(os.path.dirname(os.path.abspath(__file__)), filename)

if __name__ == '__main__':
    # Sunucu başlarken database.json'u kontrol et
    json_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'database.json')
    if not os.path.exists(json_file):
        print("📄 database.json bulunamadı, Excel'den oluşturulmaya çalışılıyor...")
        if excel_to_json():
            print("✓ database.json başarıyla oluşturuldu")
        else:
            print("⚠️ database.json oluşturulamadı - Excel dosyası işlenemedi")
    else:
        print("✓ database.json mevcut")
    
    # Scheduler'ı başlat
    if not scheduler.running:
        # Günde 3 kez çalıştır (10:30, 18:00, 23:45)
        scheduler.add_job(
            run_daily_update,
            trigger=CronTrigger(hour=10, minute=30),
            id='daily_update_morning',
            name='Sabah Güncelleme (10:30)',
            replace_existing=True
        )
        scheduler.add_job(
            run_daily_update,
            trigger=CronTrigger(hour=18, minute=0),
            id='daily_update_evening',
            name='Akşam Güncelleme (18:00)',
            replace_existing=True
        )
        scheduler.add_job(
            run_daily_update,
            trigger=CronTrigger(hour=23, minute=45),
            id='daily_update_night',
            name='Gece Güncelleme (23:45)',
            replace_existing=True
        )
        scheduler.start()
        print("✓ Scheduler başlatıldı. Günde 3 kez (10:30, 18:00, 23:45) otomatik güncelleme çalışacak.")
    
    # Flask sunucusunu başlat (port 5000)
    print("✓ Web sunucusu başladı: http://localhost:5000")
    app.run(host='0.0.0.0', port=5000, debug=True, use_reloader=False)
