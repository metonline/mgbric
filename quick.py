import openpyxl, json

wb = openpyxl.load_workbook('database.xlsx', data_only=True)
ws = wb.active
headers = [ws.cell(1, col).value for col in range(1, 9)]
data = []
for i in range(2, min(ws.max_row + 1, 102)):
    r = [ws.cell(i, j).value for j in range(1, 9)]
    if r[0]:
        data.append(dict(zip(headers, r)))
with open('database.json', 'w', encoding='utf-8') as f:
    json.dump(data, f, ensure_ascii=False, separators=(',', ':'))
