#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import openpyxl
import json
import sys

try:
    # Excel dosyasını aç
    wb = openpyxl.load_workbook('database.xlsx', data_only=True)
    ws = wb.active

    # Header'ları oku
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]

    # Tüm veriyi oku
    data = []
    for row_idx in range(2, ws.max_row + 1):
        row_data = [ws.cell(row_idx, col).value for col in range(1, ws.max_column + 1)]
        if row_data[0]:  # İlk kolon boş değilse
            record = dict(zip(headers, row_data))
            data.append(record)

    # database.json'u oluştur
    with open('database.json', 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, separators=(',', ':'))

    print(f"Başarılı! {len(data)} kayıt yazıldı.")
    sys.exit(0)

except Exception as e:
    print(f"Hata: {e}", file=sys.stderr)
    sys.exit(1)
