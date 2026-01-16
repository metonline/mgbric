#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from openpyxl import load_workbook

wb = load_workbook('database.xlsx')
ws = wb['Sonuçlar']

count = 0
for row in ws.iter_rows(min_row=2, values_only=False):
    if len(row) > 6:
        tournament = row[6].value
        if tournament and 'YILBAŞI AKŞAM' in str(tournament) and '26-12-2025 14:00' in str(tournament):
            row[6].value = tournament.replace('26-12-2025 14:00', '26-12-2025 20:00')
            count += 1

wb.save('database.xlsx')
print(f'✓ {count} kayıt güncellendi: YILBAŞI AKŞAM TURNUVASI 14:00 → 20:00')
