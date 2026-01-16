#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import sys
import traceback

try:
    import openpyxl
    import json
    
    print("✓ Modules imported")
    
    # Excel dosyasını aç
    wb = openpyxl.load_workbook('database.xlsx')
    ws = wb.active
    
    print(f"✓ Excel açıldı: {ws.max_row} satır, {ws.max_column} kolon")
    
    # Header'ları oku
    headers = [cell.value for cell in ws[1]]
    print(f"✓ Headers: {headers[:5]}")
    
    # Tüm veriyi oku
    data = []
    for row_idx in range(2, ws.max_row + 1):
        row = tuple(ws.cell(row_idx, col).value for col in range(1, ws.max_column + 1))
        if row[0]:  # Tarih kolonunun boş olup olmadığını kontrol et
            record = dict(zip(headers, row))
            data.append(record)
    
    print(f"✓ Toplam kayıt okundu: {len(data)}")
    
    # database.json'u oluştur
    with open('database.json', 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=None)
    
    print(f"✓ database.json yazıldı: {len(data)} kayıt")
    
    # Kontrol - dosya boyutu
    import os
    size = os.path.getsize('database.json')
    print(f"✓ Dosya boyutu: {size} bytes")
    
except Exception as e:
    print(f"❌ Hata: {e}")
    traceback.print_exc()
    sys.exit(1)
