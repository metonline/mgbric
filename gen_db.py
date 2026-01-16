import openpyxl
import json

wb = openpyxl.load_workbook('database.xlsx', data_only=True)
ws = wb.active
headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
data = []

for row_idx in range(2, ws.max_row + 1):
    row_data = [ws.cell(row_idx, col).value for col in range(1, ws.max_column + 1)]
    if row_data[0]:
        data.append(dict(zip(headers, row_data)))

with open('database.json', 'w', encoding='utf-8') as f:
    json.dump(data, f, ensure_ascii=False, separators=(',', ':'))

